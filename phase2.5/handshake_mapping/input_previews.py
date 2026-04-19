"""Load short previews of user-supplied input files for codegen prompts."""

from __future__ import annotations

import csv
import json
from pathlib import Path

# Keep token usage small for OpenAI codegen; match UI preview cap for spreadsheets.
_DEFAULT_ROWS = 50
_MAX_CELL = 400


def _truncate_cell(s: str, max_cell: int) -> str:
    s = s.replace("\n", "\\n").replace("\r", "")
    if len(s) <= max_cell:
        return s
    return s[:max_cell] + "…"


def preview_csv(path: Path, *, max_rows: int = _DEFAULT_ROWS, max_cell: int = _MAX_CELL, delimiter: str = ",") -> str:
    lines: list[str] = []
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for i, row in enumerate(reader):
            if i >= max_rows:
                lines.append(f"... truncated after {max_rows} data rows ...")
                break
            trimmed = [_truncate_cell(c, max_cell) for c in row]
            sep = "\t" if delimiter == "\t" else ","
            lines.append(sep.join(trimmed))
    return "\n".join(lines)


def preview_xlsx(
    path: Path,
    *,
    max_rows: int = _DEFAULT_ROWS,
    max_cell: int = _MAX_CELL,
) -> str:
    """First ``max_rows`` rows per worksheet for **all** sheets (tab-separated)."""
    from openpyxl import load_workbook

    parts: list[str] = []
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        parts.append(f"### Workbook sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")
        parts.append("")
        for ws in wb.worksheets:
            parts.append(f"### Sheet: {ws.title}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    parts.append(f"... truncated after {max_rows} rows in this sheet ...")
                    break
                cells = [_truncate_cell("" if c is None else str(c), max_cell) for c in row]
                parts.append("\t".join(cells))
            parts.append("")
    finally:
        wb.close()
    out = "\n".join(parts).rstrip()
    _MAX_PREVIEW_CHARS = 12_000
    if len(out) > _MAX_PREVIEW_CHARS:
        return out[:_MAX_PREVIEW_CHARS] + "\n... [total spreadsheet preview truncated for API size] ..."
    return out


def preview_text(path: Path, *, max_chars: int = 8000) -> str:
    """Read only the first chunk of bytes so huge files are not fully loaded."""
    cap = max_chars * 4 + 4
    with path.open("rb") as f:
        raw_bytes = f.read(cap)
    raw = raw_bytes.decode("utf-8", errors="replace")
    if len(raw) > max_chars:
        return raw[:max_chars] + "\n... truncated ..."
    return raw


def preview_path(path: Path) -> str:
    suf = path.suffix.lower()
    if suf == ".csv":
        return preview_csv(path)
    if suf == ".tsv":
        return preview_csv(path, delimiter="\t")
    if suf in (".xlsx", ".xlsm"):
        try:
            return preview_xlsx(path)
        except Exception as e:
            return f"<error reading {path}: {e}>"
    if suf in (".json", ".jsonl"):
        try:
            return preview_text(path, max_chars=12000)
        except OSError as e:
            return f"<error reading {path}: {e}>"
    return preview_text(path, max_chars=8000)


def build_inputs_section(paths: list[Path], *, max_total_chars: int = 14_000) -> str:
    """Cap total preview size so codegen stays under OpenAI TPM limits."""
    parts: list[str] = []
    for p in paths:
        rp = p.resolve()
        if not rp.is_file():
            parts.append(f"## {rp}\n<MISSING FILE>\n")
            continue
        try:
            body = preview_path(rp)
        except OSError as e:
            parts.append(f"## {rp}\n<error: {e}>\n")
            continue
        parts.append(f"## {rp}\n```\n{body}\n```\n")
    out = "\n".join(parts) if parts else "(no input files provided)\n"
    if len(out) > max_total_chars:
        return (
            out[:max_total_chars]
            + "\n... [input previews truncated: too many/large uploads for one codegen request] ..."
        )
    return out


def phase2_columns_snippets(phase2_output: Path, table_slugs: list[str]) -> str:
    """Include Phase 2 columns.json for each table so codegen sees dtypes/descriptions."""
    parts: list[str] = []
    root = phase2_output / "tables"
    for slug in table_slugs:
        cj = root / slug / "columns.json"
        if not cj.is_file():
            parts.append(f"### {slug}\n<no columns.json at {cj}>\n")
            continue
        try:
            data = json.loads(cj.read_text(encoding="utf-8"))
            parts.append(f"### {slug}\n```json\n{json.dumps(data, indent=2)[:5000]}\n```\n")
        except (OSError, json.JSONDecodeError) as e:
            parts.append(f"### {slug}\n<error reading columns.json: {e}>\n")
    return "\n".join(parts)
