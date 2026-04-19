"""Agent tools.

Every tool is a plain Python function. ``build_tool_specs`` converts them
into OpenAI tool-call schemas. All filesystem access is restricted to the
configured output directory plus a per-session upload subdirectory.
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import httpx
import pandas as pd
from bs4 import BeautifulSoup

from . import git_ops
from .settings import get_settings

log = logging.getLogger(__name__)

MAX_BYTES = 256_000
MAX_FETCH_BYTES = 500_000


class ToolError(RuntimeError):
    pass


# ---------------------------------------------------------------------------
# Path safety


def _resolve_safe(path: str, session_id: str | None = None) -> Path:
    """Resolve ``path`` inside one of the allow-listed roots.

    ``output/...`` and bare relative paths land under the output dir.
    ``uploads/...`` lands under the session upload dir.
    Absolute paths and ``..`` traversal are rejected.
    """
    settings = get_settings()
    raw = path.strip()
    if not raw:
        raise ToolError("path is required")
    if raw.startswith("/"):
        raise ToolError("absolute paths are not allowed")

    parts = raw.replace("\\", "/").split("/")
    if any(p in ("..",) for p in parts):
        raise ToolError("path traversal is not allowed")

    root: Path
    rel_parts: list[str]
    if parts[0] == "uploads":
        if not session_id:
            raise ToolError("uploads/ requires an active session")
        root = settings.upload_path / session_id
        rel_parts = parts[1:]
    elif parts[0] == "output":
        root = settings.output_path
        rel_parts = parts[1:]
    else:
        root = settings.output_path
        rel_parts = parts

    resolved = (root / Path(*rel_parts)).resolve()
    try:
        resolved.relative_to(root.resolve())
    except ValueError as exc:
        raise ToolError("path escapes allowed root") from exc
    return resolved


# ---------------------------------------------------------------------------
# Tool context


@dataclass
class ToolContext:
    session_id: str
    written_paths: set[Path]

    def __init__(self, session_id: str) -> None:
        self.session_id = session_id
        self.written_paths = set()


# ---------------------------------------------------------------------------
# File tools


def tool_list_files(ctx: ToolContext) -> str:
    settings = get_settings()
    out = settings.output_path
    up = settings.upload_path / ctx.session_id

    def _tree(root: Path, prefix: str) -> list[dict[str, Any]]:
        if not root.exists():
            return []
        entries: list[dict[str, Any]] = []
        for p in sorted(root.rglob("*")):
            if p.is_dir():
                continue
            rel = p.relative_to(root)
            entries.append(
                {
                    "path": f"{prefix}/{rel.as_posix()}",
                    "size": p.stat().st_size,
                }
            )
        return entries

    return json.dumps(
        {
            "output": _tree(out, "output"),
            "uploads": _tree(up, "uploads"),
        }
    )


def tool_read_file(ctx: ToolContext, path: str, max_bytes: int = MAX_BYTES) -> str:
    target = _resolve_safe(path, ctx.session_id)
    if not target.exists() or not target.is_file():
        raise ToolError(f"file not found: {path}")
    if target.suffix.lower() in (".xlsx", ".xlsm"):
        raise ToolError(
            "Excel workbooks are binary and too large to read as text. "
            "Call preview_excel with the same path for a capped row sample per sheet."
        )
    data = target.read_bytes()[:max_bytes]
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("utf-8", errors="replace")
    truncated = target.stat().st_size > len(data)
    return json.dumps({"path": path, "truncated": truncated, "content": text})


def tool_write_file(ctx: ToolContext, path: str, content: str) -> str:
    target = _resolve_safe(path, ctx.session_id)
    if path.startswith("uploads/"):
        raise ToolError("uploads are read-only")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(content, encoding="utf-8")
    ctx.written_paths.add(target)
    return json.dumps({"path": path, "bytes": len(content.encode("utf-8"))})


# ---------------------------------------------------------------------------
# Data preview tools


def tool_preview_csv(ctx: ToolContext, path: str, n: int = 20) -> str:
    target = _resolve_safe(path, ctx.session_id)
    if not target.exists():
        raise ToolError(f"file not found: {path}")
    n = max(1, min(int(n), 200))
    try:
        df = pd.read_csv(target, nrows=n)
    except Exception as exc:
        raise ToolError(f"failed to parse CSV: {exc}") from exc
    total_rows = sum(1 for _ in open(target, "rb")) - 1  # minus header
    summary = {
        "path": path,
        "sampled_rows": len(df),
        "total_rows_estimate": max(total_rows, len(df)),
        "columns": [
            {
                "name": str(col),
                "dtype": str(df[col].dtype),
                "non_null": int(df[col].notna().sum()),
                "null": int(df[col].isna().sum()),
                "sample_values": [
                    None if pd.isna(v) else str(v)
                    for v in df[col].head(5).tolist()
                ],
            }
            for col in df.columns
        ],
        "head": df.head(n).astype(object).where(df.head(n).notna(), None).to_dict("records"),
    }
    return json.dumps(summary, default=str)


def tool_preview_excel(
    ctx: ToolContext,
    path: str,
    rows: int = 50,
    max_cols: int = 48,
) -> str:
    """Sample **every** worksheet: first N rows each (N is lowered until JSON fits model budget)."""
    from openpyxl import load_workbook

    target = _resolve_safe(path, ctx.session_id)
    if not target.exists() or not target.is_file():
        raise ToolError(f"file not found: {path}")
    if target.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ToolError("preview_excel only supports .xlsx and .xlsm files")

    want_rows = max(1, min(int(rows), 100))
    max_cols_arg = max(8, min(int(max_cols), 80))
    max_cell = 120

    def _cell(v: object) -> str:
        if v is None:
            return ""
        s = str(v).replace("\n", " ").replace("\r", "")
        return s if len(s) <= max_cell else s[:max_cell] + "…"

    def _build_payload(wb: Any, per_sheet_rows: int, cols_cap: int) -> dict[str, Any]:
        out_sheets: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            grid: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= per_sheet_rows:
                    break
                vals = list(row)[:cols_cap]
                grid.append([_cell(c) for c in vals])
            out_sheets.append(
                {
                    "sheet": ws.title,
                    "rows_in_preview": len(grid),
                    "truncated_per_sheet": len(grid) >= per_sheet_rows,
                    "row_samples": grid,
                }
            )
        return {
            "path": path,
            "sheet_names": list(wb.sheetnames),
            "sheet_count": len(wb.sheetnames),
            "note": (
                f"All {len(wb.sheetnames)} sheet(s) sampled; up to {per_sheet_rows} row(s) each, "
                f"{cols_cap} column(s) per row — not the full workbook."
            ),
            "sheets": out_sheets,
        }

    _MAX_JSON = 19_000
    row_candidates = sorted({want_rows, 35, 25, 18, 12, 8, 5, 3, 2, 1}, reverse=True)
    col_candidates = [max_cols_arg, 40, 32, 24, 16, 12, 8]

    wb = load_workbook(filename=target, read_only=True, data_only=True)
    try:
        for cols_cap in col_candidates:
            for per_sheet_rows in row_candidates:
                payload = _build_payload(wb, per_sheet_rows, cols_cap)
                raw = json.dumps(payload, default=str)
                if len(raw) <= _MAX_JSON:
                    return raw
            payload = _build_payload(wb, 1, cols_cap)
            raw = json.dumps(payload, default=str)
            if len(raw) <= _MAX_JSON:
                return raw
        return json.dumps(
            {
                "path": path,
                "sheet_names": list(wb.sheetnames),
                "sheet_count": len(wb.sheetnames),
                "error": (
                    "Row samples still too large after shrinking rows/columns; "
                    "sheet list is complete. Try CSV export or a narrower workbook."
                ),
            },
            default=str,
        )
    finally:
        wb.close()


def tool_preview_json(ctx: ToolContext, path: str) -> str:
    target = _resolve_safe(path, ctx.session_id)
    if not target.exists():
        raise ToolError(f"file not found: {path}")
    text = target.read_text(encoding="utf-8", errors="replace")
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ToolError(f"not valid JSON: {exc}") from exc

    def _shape(value: Any, depth: int = 0) -> Any:
        if depth > 3:
            return "..."
        if isinstance(value, dict):
            return {k: _shape(v, depth + 1) for k, v in list(value.items())[:20]}
        if isinstance(value, list):
            return [_shape(v, depth + 1) for v in value[:5]]
        return type(value).__name__

    sample = data
    if isinstance(data, list):
        sample = data[:3]
    return json.dumps(
        {
            "path": path,
            "top_level_type": type(data).__name__,
            "length": len(data) if isinstance(data, (list, dict)) else None,
            "shape": _shape(data),
            "sample": sample,
        },
        default=str,
    )


# ---------------------------------------------------------------------------
# Network tools


def tool_fetch_url(ctx: ToolContext, url: str) -> str:
    if not url.startswith(("http://", "https://")):
        raise ToolError("url must start with http:// or https://")
    try:
        with httpx.Client(follow_redirects=True, timeout=20.0) as client:
            resp = client.get(url, headers={"user-agent": "phase2-agent/0.1"})
    except httpx.HTTPError as exc:
        raise ToolError(f"fetch failed: {exc}") from exc

    ctype = resp.headers.get("content-type", "")
    body = resp.content[:MAX_FETCH_BYTES]
    if "html" in ctype:
        soup = BeautifulSoup(body, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = soup.get_text("\n", strip=True)
    else:
        text = body.decode("utf-8", errors="replace")

    return json.dumps(
        {
            "url": url,
            "status": resp.status_code,
            "content_type": ctype,
            "text": text[:MAX_FETCH_BYTES],
            "truncated": len(resp.content) > MAX_FETCH_BYTES,
        }
    )


def tool_call_api(
    ctx: ToolContext,
    method: str,
    url: str,
    headers: dict[str, str] | None = None,
    params: dict[str, Any] | None = None,
    json_body: dict[str, Any] | None = None,
) -> str:
    if not url.startswith(("http://", "https://")):
        raise ToolError("url must start with http:// or https://")
    method = (method or "GET").upper()
    try:
        with httpx.Client(follow_redirects=True, timeout=30.0) as client:
            resp = client.request(
                method,
                url,
                headers=headers or {},
                params=params or None,
                json=json_body,
            )
    except httpx.HTTPError as exc:
        raise ToolError(f"request failed: {exc}") from exc

    body = resp.content[:MAX_FETCH_BYTES]
    text = body.decode("utf-8", errors="replace")
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        parsed = None
    return json.dumps(
        {
            "status": resp.status_code,
            "headers": dict(resp.headers),
            "json": parsed,
            "text": text if parsed is None else None,
            "truncated": len(resp.content) > MAX_FETCH_BYTES,
        },
        default=str,
    )


# ---------------------------------------------------------------------------
# Git tool


def tool_git_commit(ctx: ToolContext, message: str) -> str:
    if not ctx.written_paths:
        return json.dumps({"committed": False, "reason": "no pending writes"})
    record = git_ops.commit_output(message, paths=list(ctx.written_paths))
    ctx.written_paths.clear()
    if record is None:
        return json.dumps({"committed": False, "reason": "nothing staged"})
    return json.dumps(
        {
            "committed": True,
            "sha": record.short_sha,
            "message": record.message,
            "files": record.files,
        }
    )


# ---------------------------------------------------------------------------
# Dispatch table & OpenAI schemas


TOOL_IMPLS: dict[str, Callable[..., str]] = {
    "list_files": tool_list_files,
    "read_file": tool_read_file,
    "write_file": tool_write_file,
    "preview_csv": tool_preview_csv,
    "preview_excel": tool_preview_excel,
    "preview_json": tool_preview_json,
    "fetch_url": tool_fetch_url,
    "call_api": tool_call_api,
    "git_commit": tool_git_commit,
}


def build_tool_specs() -> list[dict[str, Any]]:
    return [
        {
            "type": "function",
            "function": {
                "name": "list_files",
                "description": (
                    "List files in the agent's output directory and the current "
                    "session's uploads directory."
                ),
                "parameters": {"type": "object", "properties": {}, "required": []},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_file",
                "description": (
                    "Read a text file. Paths starting with 'uploads/' read the "
                    "user's uploaded files; any other relative path is rooted at "
                    "phase2/output/. Do not use for .xlsx/.xlsm — use preview_excel."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "max_bytes": {"type": "integer", "default": MAX_BYTES},
                    },
                    "required": ["path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "write_file",
                "description": (
                    "Write a text file under phase2/output/. Use this to create "
                    "or update structured table descriptions and column-info "
                    "artifacts. Overwrites any existing file."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Relative path under output/, e.g. 'tables/invoices/description.md'.",
                        },
                        "content": {"type": "string"},
                    },
                    "required": ["path", "content"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "preview_csv",
                "description": "Summarize the first N rows of a CSV file with column dtypes and null counts.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "n": {"type": "integer", "default": 20},
                    },
                    "required": ["path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "preview_excel",
                "description": (
                    "Sample **every worksheet** in an .xlsx/.xlsm: first N rows per sheet "
                    "(N may be lowered automatically so the result fits the model). "
                    "Response includes full sheet_names. Never use read_file on Excel."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "rows": {"type": "integer", "default": 50},
                        "max_cols": {"type": "integer", "default": 48},
                    },
                    "required": ["path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "preview_json",
                "description": "Summarize a JSON file's shape and a small sample of its content.",
                "parameters": {
                    "type": "object",
                    "properties": {"path": {"type": "string"}},
                    "required": ["path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "fetch_url",
                "description": (
                    "HTTP GET a public URL and return readable text (HTML is "
                    "stripped). Use for documentation pages."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {"url": {"type": "string"}},
                    "required": ["url"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "call_api",
                "description": (
                    "Call an arbitrary HTTP API. Use this when the user has "
                    "supplied an API endpoint plus credentials to explore."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "method": {"type": "string", "default": "GET"},
                        "url": {"type": "string"},
                        "headers": {"type": "object", "additionalProperties": {"type": "string"}},
                        "params": {"type": "object"},
                        "json_body": {"type": "object"},
                    },
                    "required": ["url"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "git_commit",
                "description": (
                    "Commit any files you've written this turn to the repo. "
                    "Scoped to phase2/output/. Call this after a logical unit of "
                    "work (e.g. finishing description of one table)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {"message": {"type": "string"}},
                    "required": ["message"],
                },
            },
        },
    ]


def dispatch(ctx: ToolContext, name: str, arguments: dict[str, Any]) -> str:
    impl = TOOL_IMPLS.get(name)
    if impl is None:
        raise ToolError(f"unknown tool: {name}")
    try:
        return impl(ctx, **arguments)
    except ToolError:
        raise
    except TypeError as exc:
        raise ToolError(f"bad arguments for {name}: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        log.exception("tool %s failed", name)
        raise ToolError(f"{name} failed: {exc}") from exc
