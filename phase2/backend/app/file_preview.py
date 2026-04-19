"""Build JSON payloads for file preview (text, binary, or capped XLSX sheet preview)."""

from __future__ import annotations

from pathlib import Path

# First N rows per worksheet for XLSX upload/preview (avoid loading huge sheets).
XLSX_PREVIEW_ROWS = 50


def content_payload(logical_path: str, target: Path) -> dict:
    """Return a dict suitable for ``GET`` upload/handshake content responses."""
    suffix = target.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _xlsx_payload(logical_path, target, XLSX_PREVIEW_ROWS)

    data = target.read_bytes()
    try:
        text = data.decode("utf-8")
        binary = False
    except UnicodeDecodeError:
        text = ""
        binary = True
    return {
        "path": logical_path,
        "size": len(data),
        "mtime": target.stat().st_mtime,
        "binary": binary,
        "content": text,
    }


def _xlsx_payload(logical_path: str, target: Path, max_rows: int) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        stat = target.stat()
        return {
            "path": logical_path,
            "size": stat.st_size,
            "mtime": stat.st_mtime,
            "binary": False,
            "content": (
                "Excel preview unavailable: openpyxl is not installed in this API environment.\n\n"
                "Fix: cd phase2/backend && . .venv/bin/activate && pip install -r requirements.txt\n"
                "Then restart the backend (make dev)."
            ),
        }

    sheets_out: list[dict] = []
    wb = load_workbook(filename=target, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                rows.append(["" if c is None else str(c) for c in row])
            truncated = len(rows) >= max_rows
            sheets_out.append(
                {
                    "name": ws.title,
                    "rows": rows,
                    "truncated": truncated,
                }
            )
    finally:
        wb.close()

    stat = target.stat()
    return {
        "path": logical_path,
        "size": stat.st_size,
        "mtime": stat.st_mtime,
        "binary": False,
        "content": "",
        "xlsx_preview": {
            "max_rows_per_sheet": max_rows,
            "sheets": sheets_out,
        },
    }
